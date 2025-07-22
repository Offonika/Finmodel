# import_ozon_realization_v2_raw.py
# -------------------------------------------------------------------
# Скачивает отчет Ozon /v2/finance/realization и сохраняет в Excel
# Поля переводятся на русский, файл создаётся отдельно
# -------------------------------------------------------------------

import os
import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR  = os.path.join(BASE_DIR, "reports")
os.makedirs(REPORT_DIR, exist_ok=True)

def fetch_raw_realization(client_id, api_key, year, month):
    url = "https://api-seller.ozon.ru/v2/finance/realization"
    headers = {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json"
    }
    payload = {
        "month": month,
        "year": year
    }

    print(f"→ Запрос данных за {month:02d}.{year}")
    response = requests.post(url, json=payload, headers=headers)
    response.raise_for_status()
    return response.json()["result"]["rows"]

def flatten_rows(rows):
    flat = []
    for r in rows:
        row = {
            "rowNumber": r.get("rowNumber"),
            "commission_ratio": r.get("commission_ratio"),
            "seller_price_per_instance": r.get("seller_price_per_instance")
        }

        item = r.get("item") or {}
        row.update({f"item_{k}": v for k, v in item.items()})

        delivery = r.get("delivery_commission") or {}
        row.update({f"delivery_{k}": v for k, v in delivery.items()})

        returns = r.get("return_commission") or {}
        row.update({f"return_{k}": v for k, v in returns.items()})

        flat.append(row)
    return pd.DataFrame(flat)

def translate_columns(df: pd.DataFrame):
    df.rename(columns={
        "rowNumber": "Номер строки",
        "commission_ratio": "Комиссия (%)",
        "seller_price_per_instance": "Цена продавца за единицу",
        "item_sku": "SKU товара",
        "item_offer_id": "Артикул продавца",
        "item_barcode": "Штрихкод",
        "item_name": "Наименование товара",
        "delivery_amount": "Доставка — Сумма начислений",
        "delivery_bonus": "Доставка — Бонус",
        "delivery_commission": "Доставка — Комиссия",
        "delivery_compensation": "Доставка — Компенсация",
        "delivery_price_per_instance": "Доставка — Цена за единицу",
        "delivery_quantity": "Доставка — Кол-во",
        "delivery_standard_fee": "Доставка — Базовый тариф",
        "delivery_bank_coinvestment": "Доставка — Коинвест. банка",
        "delivery_stars": "Доставка — Баллы",
        "delivery_pick_up_point_coinvestment": "Доставка — Коинвест. ПВЗ",
        "delivery_total": "Доставка — Всего",
        "return_amount": "Возврат — Сумма начислений",
        "return_bonus": "Возврат — Бонус",
        "return_commission": "Возврат — Комиссия",
        "return_compensation": "Возврат — Компенсация",
        "return_price_per_instance": "Возврат — Цена за единицу",
        "return_quantity": "Возврат — Кол-во",
        "return_standard_fee": "Возврат — Базовый тариф",
        "return_bank_coinvestment": "Возврат — Коинвест. банка",
        "return_stars": "Возврат — Баллы",
        "return_pick_up_point_coinvestment": "Возврат — Коинвест. ПВЗ",
        "return_total": "Возврат — Всего"
    }, inplace=True)
    return df

def save_to_excel(df: pd.DataFrame, filename: str, sheet_name: str):
    path = os.path.join(REPORT_DIR, filename)
    df.to_excel(path, index=False, sheet_name=sheet_name)

    # Автоширина + жирные заголовки
    wb = load_workbook(path)
    ws = wb[sheet_name]
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)
        col[0].font = Font(bold=True)
    wb.save(path)

    print(f"✓ Сохранено в файл: {path}")

def main():
    print("=== Старт import_ozon_realization_v2_raw ===")

    # 🔐 Настройки подключения и периода:
    CLIENT_ID = "551025"
    API_KEY   = "52ba5ea1-7e09-406b-82c2-7fef7effdb81"
    YEAR  = 2025
    MONTH = 1

    try:
        rows = fetch_raw_realization(CLIENT_ID, API_KEY, YEAR, MONTH)
        if not rows:
            print("⚠️ Данных нет")
            return
    except Exception as e:
        print(f"❌ Ошибка запроса: {e}")
        return

    df = flatten_rows(rows)
    df = translate_columns(df)

    filename = f"Озон_Реализация_{YEAR}_{MONTH:02d}.xlsx"
    sheet_name = "Реализация"

    save_to_excel(df, filename, sheet_name)

if __name__ == "__main__":
    main()
